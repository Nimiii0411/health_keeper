import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import csv

def create_comprehensive_test_xlsx():
    # Tạo workbook
    wb = Workbook()
    
    # Xóa sheet mặc định
    wb.remove(wb.active)
    
    # Login Test Cases (30 test cases)
    login_tests = [
        [1, "REQ-LOGIN-001", "TC-LOGIN-001", "Valid Login with Correct Credentials", 5, "Functional", "Login", 
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email hợp lệ (user@gmail.com)\n4. Nhập mật khẩu đúng\n5. Nhấn nút 'Đăng nhập'", 
         "Hệ thống đăng nhập thành công và chuyển đến màn hình chính", "High", "No", "QA Team", "Test cơ bản cho chức năng đăng nhập"],
        
        [2, "REQ-LOGIN-002", "TC-LOGIN-002", "Invalid Login with Wrong Email", 3, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email không hợp lệ (wrongemail@gmail.com)\n4. Nhập mật khẩu bất kỳ\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống hiển thị thông báo lỗi 'Email hoặc mật khẩu không đúng'", "High", "No", "QA Team", "Test với email sai"],
         
        [3, "REQ-LOGIN-003", "TC-LOGIN-003", "Invalid Login with Wrong Password", 3, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email hợp lệ\n4. Nhập mật khẩu sai\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống hiển thị thông báo lỗi 'Email hoặc mật khẩu không đúng'", "High", "No", "QA Team", "Test với mật khẩu sai"],
         
        [4, "REQ-LOGIN-004", "TC-LOGIN-004", "Login with Empty Fields", 2, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Để trống cả email và mật khẩu\n4. Nhấn nút 'Đăng nhập'",
         "Hệ thống hiển thị thông báo yêu cầu nhập đầy đủ thông tin", "Medium", "No", "QA Team", "Test validation cho field rỗng"],
         
        [5, "REQ-LOGIN-005", "TC-LOGIN-005", "Login with Invalid Email Format", 3, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email không đúng định dạng (invalidemail)\n4. Nhập mật khẩu bất kỳ\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống hiển thị thông báo lỗi định dạng email không hợp lệ", "Medium", "No", "QA Team", "Test validation định dạng email"],
         
        [6, "REQ-LOGIN-006", "TC-LOGIN-006", "Login with Special Characters in Password", 4, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email hợp lệ\n4. Nhập mật khẩu chứa ký tự đặc biệt (@#$%^&*)\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống xử lý bình thường và đăng nhập thành công nếu thông tin đúng", "Low", "No", "QA Team", "Test mật khẩu với ký tự đặc biệt"],
         
        [7, "REQ-LOGIN-007", "TC-LOGIN-007", "Remember Login Option", 5, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập thông tin đăng nhập hợp lệ\n4. Tích chọn 'Ghi nhớ đăng nhập'\n5. Đăng nhập thành công\n6. Đóng ứng dụng và mở lại",
         "Hệ thống tự động đăng nhập mà không yêu cầu nhập lại thông tin", "Medium", "No", "QA Team", "Test tính năng ghi nhớ đăng nhập"],
         
        [8, "REQ-LOGIN-008", "TC-LOGIN-008", "Logout Functionality", 3, "Functional", "Login",
         "1. Đăng nhập vào ứng dụng\n2. Vào menu tài khoản\n3. Nhấn nút 'Đăng xuất'\n4. Xác nhận đăng xuất",
         "Hệ thống đăng xuất và quay về màn hình đăng nhập", "High", "No", "QA Team", "Test chức năng đăng xuất"],
         
        [9, "REQ-LOGIN-009", "TC-LOGIN-009", "Login with SQL Injection Attempt", 5, "Security", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email với SQL injection (admin'--)\n4. Nhập mật khẩu bất kỳ\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống từ chối đăng nhập và hiển thị thông báo lỗi", "High", "No", "QA Team", "Test bảo mật SQL injection"],
         
        [10, "REQ-LOGIN-010", "TC-LOGIN-010", "Login with Very Long Email", 4, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email rất dài (>100 ký tự)\n4. Nhập mật khẩu\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống xử lý hợp lý hoặc hiển thị thông báo giới hạn độ dài", "Low", "No", "QA Team", "Test với email quá dài"],
         
        [11, "REQ-LOGIN-011", "TC-LOGIN-011", "Login with Very Long Password", 4, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng nhập'\n3. Nhập email hợp lệ\n4. Nhập mật khẩu rất dài (>100 ký tự)\n5. Nhấn nút 'Đăng nhập'",
         "Hệ thống xử lý hợp lý hoặc hiển thị thông báo giới hạn độ dài", "Low", "No", "QA Team", "Test với mật khẩu quá dài"],
         
        [12, "REQ-LOGIN-012", "TC-LOGIN-012", "Multiple Failed Login Attempts", 8, "Security", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Thực hiện 5 lần đăng nhập sai liên tiếp\n3. Thử đăng nhập lần thứ 6",
         "Hệ thống khóa tạm thời tài khoản hoặc yêu cầu captcha", "High", "No", "QA Team", "Test bảo mật chống brute force"],
         
        [13, "REQ-LOGIN-013", "TC-LOGIN-013", "Login with Caps Lock On", 3, "Usability", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Bật Caps Lock\n3. Nhập thông tin đăng nhập\n4. Nhấn nút 'Đăng nhập'",
         "Hệ thống hiển thị cảnh báo Caps Lock hoặc xử lý phù hợp", "Low", "No", "QA Team", "Test với Caps Lock"],
         
        [14, "REQ-LOGIN-014", "TC-LOGIN-014", "Login with Copy-Paste Password", 3, "Functional", "Login",
         "1. Mở ứng dụng HealthKeeper\n2. Copy mật khẩu từ nguồn khác\n3. Paste vào trường mật khẩu\n4. Nhấn đăng nhập",
         "Hệ thống chấp nhận mật khẩu paste và đăng nhập thành công", "Medium", "No", "QA Team", "Test copy-paste mật khẩu"],
         
        [15, "REQ-LOGIN-015", "TC-LOGIN-015", "Login Screen UI Responsiveness", 5, "UI/UX", "Login",
         "1. Mở ứng dụng trên các thiết bị khác nhau\n2. Kiểm tra giao diện đăng nhập\n3. Test trên màn hình ngang/dọc",
         "Giao diện đăng nhập hiển thị đúng trên mọi thiết bị", "Medium", "No", "QA Team", "Test responsive UI"],
         
        [16, "REQ-LOGIN-016", "TC-LOGIN-016", "Login with Weak Internet Connection", 6, "Performance", "Login",
         "1. Kết nối mạng yếu (2G/Edge)\n2. Mở ứng dụng HealthKeeper\n3. Thực hiện đăng nhập với thông tin đúng",
         "Hệ thống đăng nhập thành công hoặc hiển thị thông báo kết nối", "Medium", "No", "QA Team", "Test với mạng yếu"],
         
        [17, "REQ-LOGIN-017", "TC-LOGIN-017", "Login Without Internet Connection", 4, "Functional", "Login",
         "1. Tắt kết nối internet\n2. Mở ứng dụng HealthKeeper\n3. Thử đăng nhập",
         "Hệ thống hiển thị thông báo lỗi kết nối mạng", "High", "No", "QA Team", "Test offline"],
         
        [18, "REQ-LOGIN-018", "TC-LOGIN-018", "Login with Auto-fill Credentials", 4, "Functional", "Login",
         "1. Lưu thông tin đăng nhập trong trình duyệt/app\n2. Mở ứng dụng\n3. Kiểm tra auto-fill",
         "Hệ thống tự động điền thông tin đã lưu", "Low", "No", "QA Team", "Test auto-fill"],
         
        [19, "REQ-LOGIN-019", "TC-LOGIN-019", "Login Field Validation Messages", 3, "UI/UX", "Login",
         "1. Nhập email sai format\n2. Để trống mật khẩu\n3. Kiểm tra thông báo lỗi",
         "Hiển thị thông báo lỗi rõ ràng và hữu ích", "Medium", "No", "QA Team", "Test validation messages"],
         
        [20, "REQ-LOGIN-020", "TC-LOGIN-020", "Login Button State Changes", 3, "UI/UX", "Login",
         "1. Quan sát trạng thái nút đăng nhập\n2. Nhấn đăng nhập\n3. Kiểm tra loading state",
         "Nút đăng nhập hiển thị trạng thái loading khi xử lý", "Low", "No", "QA Team", "Test button states"],
         
        [21, "REQ-LOGIN-021", "TC-LOGIN-021", "Login with Biometric Authentication", 6, "Functional", "Login",
         "1. Kích hoạt đăng nhập sinh trắc học\n2. Mở ứng dụng\n3. Sử dụng vân tay/khuôn mặt để đăng nhập",
         "Hệ thống đăng nhập thành công qua sinh trắc học", "Medium", "No", "QA Team", "Test biometric login"],
         
        [22, "REQ-LOGIN-022", "TC-LOGIN-022", "Login Session Timeout", 10, "Security", "Login",
         "1. Đăng nhập thành công\n2. Để ứng dụng không hoạt động trong thời gian dài\n3. Thử sử dụng lại",
         "Hệ thống yêu cầu đăng nhập lại sau timeout", "High", "No", "QA Team", "Test session timeout"],
         
        [23, "REQ-LOGIN-023", "TC-LOGIN-023", "Login with Unicode Characters", 4, "Functional", "Login",
         "1. Nhập email/password chứa ký tự Unicode\n2. Thực hiện đăng nhập",
         "Hệ thống xử lý đúng ký tự Unicode", "Low", "No", "QA Team", "Test Unicode support"],
         
        [24, "REQ-LOGIN-024", "TC-LOGIN-024", "Login Error Recovery", 5, "Functional", "Login",
         "1. Gây ra lỗi đăng nhập\n2. Sửa lỗi và thử lại\n3. Kiểm tra khôi phục",
         "Hệ thống cho phép thử lại sau lỗi", "Medium", "No", "QA Team", "Test error recovery"],
         
        [25, "REQ-LOGIN-025", "TC-LOGIN-025", "Login with Different Device Orientations", 4, "UI/UX", "Login",
         "1. Test đăng nhập ở chế độ dọc\n2. Xoay ngang\n3. Kiểm tra giao diện",
         "Giao diện đăng nhập hoạt động tốt ở mọi hướng", "Low", "No", "QA Team", "Test orientation"],
         
        [26, "REQ-LOGIN-026", "TC-LOGIN-026", "Login Accessibility Features", 6, "Accessibility", "Login",
         "1. Bật các tính năng hỗ trợ (TalkBack, VoiceOver)\n2. Thực hiện đăng nhập",
         "Ứng dụng hỗ trợ đầy đủ các tính năng trợ năng", "Medium", "No", "QA Team", "Test accessibility"],
         
        [27, "REQ-LOGIN-027", "TC-LOGIN-027", "Login Password Visibility Toggle", 3, "UI/UX", "Login",
         "1. Nhập mật khẩu\n2. Nhấn nút hiện/ẩn mật khẩu\n3. Kiểm tra chức năng",
         "Có thể hiện/ẩn mật khẩu khi nhập", "Low", "No", "QA Team", "Test password visibility"],
         
        [28, "REQ-LOGIN-028", "TC-LOGIN-028", "Login with Virtual Keyboard", 4, "UI/UX", "Login",
         "1. Sử dụng bàn phím ảo trên thiết bị\n2. Nhập thông tin đăng nhập\n3. Kiểm tra trải nghiệm",
         "Bàn phím ảo hoạt động mượt mà với form đăng nhập", "Low", "No", "QA Team", "Test virtual keyboard"],
         
        [29, "REQ-LOGIN-029", "TC-LOGIN-029", "Login Performance Under Load", 15, "Performance", "Login",
         "1. Thực hiện nhiều request đăng nhập cùng lúc\n2. Đo thời gian phản hồi",
         "Hệ thống xử lý đăng nhập trong thời gian hợp lý dưới tải cao", "Medium", "No", "QA Team", "Test performance"],
         
        [30, "REQ-LOGIN-030", "TC-LOGIN-030", "Login Data Encryption Verification", 8, "Security", "Login",
         "1. Thực hiện đăng nhập\n2. Kiểm tra dữ liệu truyền qua network\n3. Xác minh mã hóa",
         "Thông tin đăng nhập được mã hóa khi truyền", "High", "No", "QA Team", "Test data encryption"]
    ]
    
    # Register Test Cases (30 test cases) 
    register_tests = [
        [1, "REQ-REG-001", "TC-REG-001", "Valid Registration with All Required Fields", 8, "Functional", "Register",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn vào nút 'Đăng ký'\n3. Nhập họ tên hợp lệ\n4. Nhập email hợp lệ chưa sử dụng\n5. Nhập mật khẩu mạnh\n6. Xác nhận mật khẩu\n7. Nhấn 'Đăng ký'",
         "Hệ thống tạo tài khoản thành công và chuyển đến màn hình chính", "High", "No", "QA Team", "Test đăng ký cơ bản"],
         
        [2, "REQ-REG-002", "TC-REG-002", "Registration with Existing Email", 5, "Functional", "Register",
         "1. Mở ứng dụng HealthKeeper\n2. Nhấn 'Đăng ký'\n3. Nhập thông tin hợp lệ\n4. Sử dụng email đã tồn tại\n5. Nhấn 'Đăng ký'",
         "Hệ thống hiển thị thông báo 'Email đã được sử dụng'", "High", "No", "QA Team", "Test email trùng"],
         
        [3, "REQ-REG-003", "TC-REG-003", "Registration with Weak Password", 4, "Functional", "Register",
         "1. Nhập thông tin đăng ký\n2. Sử dụng mật khẩu yếu (ít hơn 8 ký tự)\n3. Nhấn đăng ký",
         "Hệ thống yêu cầu mật khẩu mạnh hơn", "Medium", "No", "QA Team", "Test mật khẩu yếu"],
         
        [4, "REQ-REG-004", "TC-REG-004", "Registration with Mismatched Password", 3, "Functional", "Register",
         "1. Nhập thông tin đăng ký\n2. Nhập mật khẩu\n3. Nhập xác nhận mật khẩu khác\n4. Nhấn đăng ký",
         "Hệ thống hiển thị lỗi 'Mật khẩu xác nhận không khớp'", "Medium", "No", "QA Team", "Test mật khẩu không khớp"],
         
        [5, "REQ-REG-005", "TC-REG-005", "Registration with Empty Required Fields", 3, "Functional", "Register",
         "1. Mở form đăng ký\n2. Để trống các trường bắt buộc\n3. Nhấn đăng ký",
         "Hệ thống hiển thị thông báo yêu cầu nhập đầy đủ", "Medium", "No", "QA Team", "Test field bắt buộc"],
         
        [6, "REQ-REG-006", "TC-REG-006", "Registration with Invalid Email Format", 3, "Functional", "Register",
         "1. Nhập thông tin đăng ký\n2. Sử dụng email sai định dạng\n3. Nhấn đăng ký",
         "Hệ thống hiển thị lỗi định dạng email", "Medium", "No", "QA Team", "Test format email"],
         
        [7, "REQ-REG-007", "TC-REG-007", "Registration with Special Characters in Name", 4, "Functional", "Register",
         "1. Nhập họ tên chứa ký tự đặc biệt\n2. Điền các thông tin khác\n3. Nhấn đăng ký",
         "Hệ thống xử lý hoặc thông báo lỗi phù hợp", "Low", "No", "QA Team", "Test ký tự đặc biệt"],
         
        [8, "REQ-REG-008", "TC-REG-008", "Registration with Numbers in Name", 3, "Functional", "Register",
         "1. Nhập họ tên chứa số\n2. Điền thông tin khác\n3. Nhấn đăng ký",
         "Hệ thống xử lý hoặc từ chối họ tên có số", "Low", "No", "QA Team", "Test số trong tên"],
         
        [9, "REQ-REG-009", "TC-REG-009", "Registration with Very Long Name", 4, "Functional", "Register",
         "1. Nhập họ tên rất dài (>100 ký tự)\n2. Điền thông tin khác\n3. Nhấn đăng ký",
         "Hệ thống giới hạn độ dài hoặc thông báo lỗi", "Low", "No", "QA Team", "Test tên quá dài"],
         
        [10, "REQ-REG-010", "TC-REG-010", "Registration Password Strength Validation", 5, "Security", "Register",
         "1. Thử các loại mật khẩu khác nhau\n2. Kiểm tra yêu cầu độ mạnh\n3. Quan sát phản hồi hệ thống",
         "Hệ thống yêu cầu mật khẩu đủ mạnh theo tiêu chuẩn", "High", "No", "QA Team", "Test độ mạnh password"],
         
        [11, "REQ-REG-011", "TC-REG-011", "Registration with Common Passwords", 4, "Security", "Register",
         "1. Thử đăng ký với mật khẩu phổ biến (123456, password)\n2. Nhấn đăng ký",
         "Hệ thống từ chối mật khẩu phổ biến", "Medium", "No", "QA Team", "Test mật khẩu phổ biến"],
         
        [12, "REQ-REG-012", "TC-REG-012", "Registration Email Format Validation", 5, "Functional", "Register",
         "1. Thử nhiều định dạng email khác nhau\n2. Kiểm tra validation\n3. Quan sát phản hồi",
         "Hệ thống chỉ chấp nhận email đúng định dạng", "Medium", "No", "QA Team", "Test validation email"],
         
        [13, "REQ-REG-013", "TC-REG-013", "Registration Terms and Conditions", 4, "Legal", "Register",
         "1. Kiểm tra checkbox điều khoản\n2. Thử đăng ký không tích chọn\n3. Đọc nội dung điều khoản",
         "Phải đồng ý điều khoản mới được đăng ký", "High", "No", "QA Team", "Test điều khoản"],
         
        [14, "REQ-REG-014", "TC-REG-014", "Registration Email Verification", 8, "Functional", "Register",
         "1. Hoàn tất đăng ký\n2. Kiểm tra email xác nhận\n3. Click link xác nhận",
         "Hệ thống gửi email xác nhận và kích hoạt tài khoản", "High", "No", "QA Team", "Test xác nhận email"],
         
        [15, "REQ-REG-015", "TC-REG-015", "Registration Duplicate Prevention", 6, "Functional", "Register",
         "1. Đăng ký tài khoản thành công\n2. Thử đăng ký lại với cùng thông tin",
         "Hệ thống ngăn chặn đăng ký trùng lặp", "Medium", "No", "QA Team", "Test chống trùng lặp"],
         
        [16, "REQ-REG-016", "TC-REG-016", "Registration Form Auto-Save", 5, "UX", "Register",
         "1. Điền một phần form đăng ký\n2. Thoát ứng dụng\n3. Quay lại form",
         "Form giữ lại dữ liệu đã nhập", "Low", "No", "QA Team", "Test auto-save"],
         
        [17, "REQ-REG-017", "TC-REG-017", "Registration Input Field Limits", 6, "Functional", "Register",
         "1. Test giới hạn ký tự của từng field\n2. Nhập vượt quá giới hạn\n3. Kiểm tra xử lý",
         "Mỗi field có giới hạn phù hợp và thông báo rõ ràng", "Medium", "No", "QA Team", "Test giới hạn input"],
         
        [18, "REQ-REG-018", "TC-REG-018", "Registration Gender Selection", 3, "Functional", "Register",
         "1. Kiểm tra tùy chọn giới tính\n2. Chọn các giá trị khác nhau\n3. Hoàn tất đăng ký",
         "Có thể chọn giới tính và lưu thành công", "Low", "No", "QA Team", "Test chọn giới tính"],
         
        [19, "REQ-REG-019", "TC-REG-019", "Registration Date of Birth", 5, "Functional", "Register",
         "1. Nhập ngày sinh\n2. Thử các định dạng khác nhau\n3. Kiểm tra validation tuổi",
         "Hệ thống validation ngày sinh hợp lệ và tuổi phù hợp", "Medium", "No", "QA Team", "Test ngày sinh"],
         
        [20, "REQ-REG-020", "TC-REG-020", "Registration Phone Number", 4, "Functional", "Register",
         "1. Nhập số điện thoại\n2. Thử các định dạng khác nhau\n3. Kiểm tra validation",
         "Hệ thống validation số điện thoại đúng định dạng", "Medium", "No", "QA Team", "Test số điện thoại"],
         
        [21, "REQ-REG-021", "TC-REG-021", "Registration Profile Picture Upload", 6, "Functional", "Register",
         "1. Chọn ảnh đại diện\n2. Upload ảnh\n3. Hoàn tất đăng ký",
         "Có thể upload và lưu ảnh đại diện", "Low", "No", "QA Team", "Test upload ảnh"],
         
        [22, "REQ-REG-022", "TC-REG-022", "Registration Form Validation Messages", 4, "UX", "Register",
         "1. Nhập dữ liệu sai các field\n2. Quan sát thông báo lỗi\n3. Kiểm tra độ rõ ràng",
         "Thông báo lỗi rõ ràng và hướng dẫn sửa", "Medium", "No", "QA Team", "Test thông báo lỗi"],
         
        [23, "REQ-REG-023", "TC-REG-023", "Registration Success Confirmation", 3, "UX", "Register",
         "1. Hoàn tất đăng ký thành công\n2. Kiểm tra màn hình xác nhận\n3. Kiểm tra hướng dẫn tiếp theo",
         "Hiển thị xác nhận thành công và hướng dẫn rõ ràng", "Medium", "No", "QA Team", "Test xác nhận thành công"],
         
        [24, "REQ-REG-024", "TC-REG-024", "Registration Back Button Functionality", 3, "UX", "Register",
         "1. Điền form đăng ký\n2. Nhấn nút Back\n3. Kiểm tra xử lý dữ liệu",
         "Có thể quay lại và dữ liệu được xử lý phù hợp", "Low", "No", "QA Team", "Test nút Back"],
         
        [25, "REQ-REG-025", "TC-REG-025", "Registration Network Error Handling", 6, "Functional", "Register",
         "1. Điền form đăng ký\n2. Tắt mạng khi submit\n3. Kiểm tra xử lý lỗi",
         "Hệ thống thông báo lỗi mạng và cho phép thử lại", "High", "No", "QA Team", "Test lỗi mạng"],
         
        [26, "REQ-REG-026", "TC-REG-026", "Registration Server Error Handling", 5, "Functional", "Register",
         "1. Thực hiện đăng ký khi server lỗi\n2. Kiểm tra phản hồi\n3. Test retry mechanism",
         "Xử lý lỗi server gracefully và có cơ chế retry", "Medium", "No", "QA Team", "Test lỗi server"],
         
        [27, "REQ-REG-027", "TC-REG-027", "Registration Social Media Integration", 6, "Functional", "Register",
         "1. Thử đăng ký qua Facebook/Google\n2. Cấp quyền truy cập\n3. Hoàn tất đăng ký",
         "Có thể đăng ký thành công qua social media", "Medium", "No", "QA Team", "Test social login"],
         
        [28, "REQ-REG-028", "TC-REG-028", "Registration Accessibility Support", 6, "Accessibility", "Register",
         "1. Bật tính năng trợ năng\n2. Thực hiện đăng ký\n3. Kiểm tra hỗ trợ",
         "Form đăng ký hỗ trợ đầy đủ accessibility", "Medium", "No", "QA Team", "Test accessibility"],
         
        [29, "REQ-REG-029", "TC-REG-029", "Registration Multi-language Support", 5, "Functional", "Register",
         "1. Chuyển đổi ngôn ngữ\n2. Thực hiện đăng ký\n3. Kiểm tra hiển thị",
         "Form đăng ký hiển thị đúng theo ngôn ngữ được chọn", "Low", "No", "QA Team", "Test đa ngôn ngữ"],
         
        [30, "REQ-REG-030", "TC-REG-030", "Registration Performance Under Load", 10, "Performance", "Register",
         "1. Thực hiện nhiều đăng ký cùng lúc\n2. Đo thời gian xử lý\n3. Kiểm tra hiệu suất",
         "Hệ thống xử lý đăng ký hiệu quả dưới tải cao", "Medium", "No", "QA Team", "Test performance"]
    ]
    
    # Health Diary Test Cases (30 test cases)
    diary_tests = [
        [1, "REQ-DIARY-001", "TC-DIARY-001", "Add New Health Diary Entry", 10, "Functional", "Health Diary",
         "1. Đăng nhập vào ứng dụng\n2. Vào 'Nhật ký sức khỏe'\n3. Nhấn 'Thêm mới'\n4. Nhập cân nặng, chiều cao\n5. Chọn bài tập\n6. Nhập ghi chú\n7. Lưu",
         "Hệ thống lưu thành công và hiển thị trong danh sách", "High", "No", "QA Team", "Test thêm nhật ký cơ bản"],
         
        [2, "REQ-DIARY-002", "TC-DIARY-002", "View Health Diary History", 5, "Functional", "Health Diary",
         "1. Đăng nhập\n2. Vào 'Nhật ký sức khỏe'\n3. Xem danh sách các bản ghi",
         "Hiển thị danh sách nhật ký theo thời gian", "Medium", "No", "QA Team", "Test xem lịch sử"],
         
        [3, "REQ-DIARY-003", "TC-DIARY-003", "Edit Existing Health Diary Entry", 8, "Functional", "Health Diary",
         "1. Vào nhật ký sức khỏe\n2. Chọn bản ghi\n3. Chỉnh sửa thông tin\n4. Lưu",
         "Cập nhật thông tin thành công", "High", "No", "QA Team", "Test chỉnh sửa nhật ký"],
         
        [4, "REQ-DIARY-004", "TC-DIARY-004", "Delete Health Diary Entry", 3, "Functional", "Health Diary",
         "1. Chọn bản ghi nhật ký\n2. Nhấn xóa\n3. Xác nhận xóa",
         "Xóa bản ghi thành công", "Medium", "No", "QA Team", "Test xóa nhật ký"],
         
        [5, "REQ-DIARY-005", "TC-DIARY-005", "Calculate BMI from Diary Data", 6, "Functional", "Health Diary",
         "1. Thêm bản ghi với cân nặng và chiều cao\n2. Kiểm tra tính toán BMI",
         "Hệ thống tính BMI chính xác", "High", "No", "QA Team", "Test tính BMI"],
         
        [6, "REQ-DIARY-006", "TC-DIARY-006", "Filter Diary by Date Range", 5, "Functional", "Health Diary",
         "1. Vào nhật ký\n2. Sử dụng bộ lọc ngày\n3. Chọn khoảng thời gian\n4. Áp dụng",
         "Hiển thị đúng các bản ghi trong khoảng thời gian", "Medium", "No", "QA Team", "Test lọc theo ngày"],
         
        [7, "REQ-DIARY-007", "TC-DIARY-007", "Add Diary Entry with Invalid Weight", 4, "Functional", "Health Diary",
         "1. Thêm nhật ký\n2. Nhập cân nặng không hợp lệ (âm/quá lớn)\n3. Lưu",
         "Hiển thị thông báo lỗi validation", "Medium", "No", "QA Team", "Test validation cân nặng"],
         
        [8, "REQ-DIARY-008", "TC-DIARY-008", "Add Diary Entry with Invalid Height", 4, "Functional", "Health Diary",
         "1. Thêm nhật ký\n2. Nhập chiều cao không hợp lệ\n3. Lưu",
         "Hiển thị thông báo lỗi validation", "Medium", "No", "QA Team", "Test validation chiều cao"],
         
        [9, "REQ-DIARY-009", "TC-DIARY-009", "View Health Progress Chart", 7, "Functional", "Health Diary",
         "1. Vào nhật ký\n2. Nhấn 'Xem biểu đồ tiến trình'\n3. Chọn loại biểu đồ",
         "Hiển thị biểu đồ tiến trình sức khỏe", "Medium", "No", "QA Team", "Test biểu đồ"],
         
        [10, "REQ-DIARY-010", "TC-DIARY-010", "Export Health Diary Data", 6, "Functional", "Health Diary",
         "1. Vào nhật ký\n2. Nhấn 'Xuất dữ liệu'\n3. Chọn định dạng\n4. Xác nhận",
         "Tạo file xuất thành công", "Low", "No", "QA Team", "Test xuất dữ liệu"],
         
        [11, "REQ-DIARY-011", "TC-DIARY-011", "Add Multiple Diary Entries Same Day", 6, "Functional", "Health Diary",
         "1. Thêm nhiều bản ghi cùng ngày\n2. Kiểm tra xử lý",
         "Hệ thống xử lý phù hợp nhiều bản ghi cùng ngày", "Medium", "No", "QA Team", "Test nhiều bản ghi cùng ngày"],
         
        [12, "REQ-DIARY-012", "TC-DIARY-012", "Diary Entry with Photo Attachment", 8, "Functional", "Health Diary",
         "1. Thêm nhật ký\n2. Đính kèm ảnh\n3. Lưu bản ghi",
         "Có thể đính kèm và lưu ảnh", "Low", "No", "QA Team", "Test đính kèm ảnh"],
         
        [13, "REQ-DIARY-013", "TC-DIARY-013", "Diary Search Functionality", 5, "Functional", "Health Diary",
         "1. Vào nhật ký\n2. Sử dụng tính năng tìm kiếm\n3. Nhập từ khóa",
         "Tìm kiếm và hiển thị kết quả chính xác", "Medium", "No", "QA Team", "Test tìm kiếm"],
         
        [14, "REQ-DIARY-014", "TC-DIARY-014", "Diary Entry Reminder Setting", 6, "Functional", "Health Diary",
         "1. Thiết lập nhắc nhở nhập nhật ký\n2. Kiểm tra thông báo",
         "Gửi nhắc nhở đúng thời gian", "Low", "No", "QA Team", "Test nhắc nhở"],
         
        [15, "REQ-DIARY-015", "TC-DIARY-015", "Diary Backup and Restore", 10, "Functional", "Health Diary",
         "1. Tạo backup dữ liệu nhật ký\n2. Xóa dữ liệu\n3. Restore từ backup",
         "Backup và restore hoạt động chính xác", "High", "No", "QA Team", "Test backup/restore"],
         
        [16, "REQ-DIARY-016", "TC-DIARY-016", "Diary Goal Setting and Tracking", 8, "Functional", "Health Diary",
         "1. Thiết lập mục tiêu sức khỏe\n2. Theo dõi tiến trình\n3. Kiểm tra báo cáo",
         "Thiết lập và theo dõi mục tiêu chính xác", "Medium", "No", "QA Team", "Test mục tiêu"],
         
        [17, "REQ-DIARY-017", "TC-DIARY-017", "Diary Exercise Duration Tracking", 5, "Functional", "Health Diary",
         "1. Thêm bài tập với thời gian\n2. Lưu và kiểm tra\n3. Xem tổng thời gian",
         "Theo dõi thời gian tập luyện chính xác", "Medium", "No", "QA Team", "Test thời gian tập"],
         
        [18, "REQ-DIARY-018", "TC-DIARY-018", "Diary Calorie Intake Tracking", 6, "Functional", "Health Diary",
         "1. Nhập thông tin calories\n2. Lưu bản ghi\n3. Xem báo cáo calories",
         "Theo dõi calories chính xác", "Medium", "No", "QA Team", "Test calories"],
         
        [19, "REQ-DIARY-019", "TC-DIARY-019", "Diary Water Intake Tracking", 4, "Functional", "Health Diary",
         "1. Nhập lượng nước uống\n2. Cập nhật trong ngày\n3. Xem báo cáo",
         "Theo dõi nước uống chính xác", "Low", "No", "QA Team", "Test nước uống"],
         
        [20, "REQ-DIARY-020", "TC-DIARY-020", "Diary Sleep Quality Tracking", 5, "Functional", "Health Diary",
         "1. Nhập thông tin giấc ngủ\n2. Đánh giá chất lượng\n3. Lưu bản ghi",
         "Theo dõi giấc ngủ chính xác", "Medium", "No", "QA Team", "Test giấc ngủ"],
         
        [21, "REQ-DIARY-021", "TC-DIARY-021", "Diary Mood Tracking", 4, "Functional", "Health Diary",
         "1. Chọn tâm trạng trong ngày\n2. Thêm ghi chú\n3. Lưu",
         "Theo dõi tâm trạng chính xác", "Low", "No", "QA Team", "Test tâm trạng"],
         
        [22, "REQ-DIARY-022", "TC-DIARY-022", "Diary Medication Tracking", 6, "Functional", "Health Diary",
         "1. Nhập thông tin thuốc\n2. Thiết lập lịch uống\n3. Theo dõi",
         "Theo dõi thuốc chính xác", "Medium", "No", "QA Team", "Test thuốc"],
         
        [23, "REQ-DIARY-023", "TC-DIARY-023", "Diary Symptoms Recording", 5, "Functional", "Health Diary",
         "1. Ghi lại triệu chứng\n2. Mô tả chi tiết\n3. Lưu bản ghi",
         "Ghi lại triệu chứng chính xác", "Medium", "No", "QA Team", "Test triệu chứng"],
         
        [24, "REQ-DIARY-024", "TC-DIARY-024", "Diary Sharing with Doctor", 7, "Functional", "Health Diary",
         "1. Chọn dữ liệu chia sẻ\n2. Gửi cho bác sĩ\n3. Kiểm tra quyền truy cập",
         "Chia sẻ dữ liệu với bác sĩ thành công", "Medium", "No", "QA Team", "Test chia sẻ"],
         
        [25, "REQ-DIARY-025", "TC-DIARY-025", "Diary Privacy Settings", 5, "Security", "Health Diary",
         "1. Thiết lập quyền riêng tư\n2. Kiểm tra truy cập\n3. Test bảo mật",
         "Dữ liệu được bảo vệ theo cài đặt", "High", "No", "QA Team", "Test bảo mật"],
         
        [26, "REQ-DIARY-026", "TC-DIARY-026", "Diary Offline Mode", 8, "Functional", "Health Diary",
         "1. Tắt mạng\n2. Thêm/sửa nhật ký\n3. Bật mạng và sync",
         "Hoạt động offline và sync khi có mạng", "Medium", "No", "QA Team", "Test offline"],
         
        [27, "REQ-DIARY-027", "TC-DIARY-027", "Diary Data Synchronization", 6, "Functional", "Health Diary",
         "1. Đăng nhập nhiều thiết bị\n2. Thêm dữ liệu\n3. Kiểm tra sync",
         "Dữ liệu đồng bộ chính xác giữa thiết bị", "High", "No", "QA Team", "Test đồng bộ"],
         
        [28, "REQ-DIARY-028", "TC-DIARY-028", "Diary Performance with Large Dataset", 12, "Performance", "Health Diary",
         "1. Tạo nhiều bản ghi (>1000)\n2. Test hiệu suất\n3. Đo thời gian tải",
         "Hiệu suất tốt với dữ liệu lớn", "Medium", "No", "QA Team", "Test performance"],
         
        [29, "REQ-DIARY-029", "TC-DIARY-029", "Diary Widget Integration", 5, "Functional", "Health Diary",
         "1. Thêm widget nhật ký\n2. Cập nhật từ widget\n3. Kiểm tra đồng bộ",
         "Widget hoạt động và đồng bộ chính xác", "Low", "No", "QA Team", "Test widget"],
         
        [30, "REQ-DIARY-030", "TC-DIARY-030", "Diary Voice Input Support", 6, "Functional", "Health Diary",
         "1. Sử dụng nhập liệu bằng giọng nói\n2. Kiểm tra độ chính xác\n3. Lưu bản ghi",
         "Nhập liệu giọng nói hoạt động chính xác", "Low", "No", "QA Team", "Test voice input"]
    ]
    
    # Headers
    headers = ["STT", "REQ ID", "TC ID", "Title", "Estimation (mins)", "Test Type", "Area", 
               "Procedure / Steps", "Expected Results", "Priority", "Change TC", "Author", "Remark"]
    
    # Tạo các worksheets
    # 1. Login Test Cases
    ws_login = wb.create_sheet("Login Test Cases")
    ws_login.append(headers)
    for test in login_tests:
        ws_login.append(test)
    
    # 2. Register Test Cases  
    ws_register = wb.create_sheet("Register Test Cases")
    ws_register.append(headers)
    for test in register_tests:
        ws_register.append(test)
        
    # 3. Health Diary Test Cases
    ws_diary = wb.create_sheet("Health Diary Test Cases")
    ws_diary.append(headers)
    for test in diary_tests:
        ws_diary.append(test)
    
    # Function để format worksheet
    def format_worksheet(ws, area_color):
        # Colors
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_fill = PatternFill(start_color=area_color, end_color=area_color, fill_type='solid')
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header
        for col in range(1, 14):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Column widths
        column_widths = [5, 15, 18, 30, 12, 12, 15, 50, 50, 10, 10, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Format data rows
        for row in range(2, 32):  # 30 test cases + header
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                if row % 2 == 0:
                    cell.fill = data_fill
                cell.border = thin_border
                
                # Text wrap for long columns
                if col in [8, 9, 13]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Row height
            ws.row_dimensions[row].height = 80
    
    # Format các sheets với màu khác nhau
    format_worksheet(ws_login, 'E8F5E8')      # Light Green for Login
    format_worksheet(ws_register, 'FFF2CC')   # Light Yellow for Register  
    format_worksheet(ws_diary, 'E1F5FE')      # Light Blue for Health Diary
    
    # Save file
    xlsx_path = r"d:\Code\School\Mobile\HealthKeeper\test_cases\HealthKeeper_TestCases_Complete.xlsx"
    wb.save(xlsx_path)
    
    print(f"✅ Đã tạo file XLSX thành công!")
    print(f"📁 File: {xlsx_path}")
    print(f"📊 Nội dung:")
    print(f"   - Login Test Cases: 30 test cases")
    print(f"   - Register Test Cases: 30 test cases") 
    print(f"   - Health Diary Test Cases: 30 test cases")
    print(f"   - Tổng cộng: 90 test cases")

if __name__ == "__main__":
    create_comprehensive_test_xlsx()
